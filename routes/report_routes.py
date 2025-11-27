from flask import Blueprint, request, jsonify, send_file
from flask_login import login_required, current_user
from models import db, Order, OrderItem, Customer, Product, Invoice, InvoiceItem
from datetime import datetime, timedelta
from sqlalchemy import func, extract, and_
from collections import defaultdict
from io import BytesIO
try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, PatternFill
    OPENPYXL_AVAILABLE = True
except ImportError:
    OPENPYXL_AVAILABLE = False
try:
    from reportlab.lib.pagesizes import letter, A4
    from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
    from reportlab.lib.styles import getSampleStyleSheet
    from reportlab.lib import colors
    REPORTLAB_AVAILABLE = True
except ImportError:
    REPORTLAB_AVAILABLE = False

report_bp = Blueprint('report', __name__)

@report_bp.route('/api/sales-summary', methods=['GET'])
@login_required
def sales_summary():
    """Get sales summary with revenue, orders, customers"""
    try:
        # Get date range (default: last 30 days)
        days = request.args.get('days', 30, type=int)
        start_date = datetime.now() - timedelta(days=days)
        
        # Total revenue from orders
        total_revenue = db.session.query(func.sum(Order.total_amount)).filter(
            Order.created_at >= start_date
        ).scalar() or 0.0
        
        # Total orders
        total_orders = Order.query.filter(
            Order.created_at >= start_date
        ).count()
        
        # Total customers
        total_customers = Customer.query.count()
        
        # Active customers (placed orders in period)
        active_customers = db.session.query(func.count(func.distinct(Order.customer_id))).filter(
            Order.created_at >= start_date
        ).scalar() or 0
        
        # Average order value
        avg_order_value = total_revenue / total_orders if total_orders > 0 else 0
        
        # Orders by status
        orders_by_status = db.session.query(
            Order.status,
            func.count(Order.id),
            func.sum(Order.total_amount)
        ).filter(
            Order.created_at >= start_date
        ).group_by(Order.status).all()
        
        status_breakdown = {
            status: {
                'count': count,
                'revenue': float(revenue or 0)
            }
            for status, count, revenue in orders_by_status
        }
        
        return jsonify({
            'success': True,
            'summary': {
                'total_revenue': float(total_revenue),
                'total_orders': total_orders,
                'total_customers': total_customers,
                'active_customers': active_customers,
                'avg_order_value': float(avg_order_value),
                'status_breakdown': status_breakdown
            }
        })
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)}), 500

@report_bp.route('/api/sales-trends', methods=['GET'])
@login_required
def sales_trends():
    """Get sales trends over time (daily, weekly, monthly)"""
    try:
        period = request.args.get('period', 'daily')  # daily, weekly, monthly
        days = request.args.get('days', 30, type=int)
        start_date = datetime.now() - timedelta(days=days)
        
        query = Order.query.filter(Order.created_at >= start_date)
        
        if period == 'daily':
            # Group by day
            trends = db.session.query(
                func.date(Order.created_at).label('date'),
                func.count(Order.id).label('orders'),
                func.sum(Order.total_amount).label('revenue')
            ).filter(
                Order.created_at >= start_date
            ).group_by(func.date(Order.created_at)).order_by('date').all()
            
            data = [{
                'date': str(trend.date),
                'orders': trend.orders,
                'revenue': float(trend.revenue or 0)
            } for trend in trends]
            
        elif period == 'weekly':
            # Group by week
            trends = db.session.query(
                extract('year', Order.created_at).label('year'),
                extract('week', Order.created_at).label('week'),
                func.count(Order.id).label('orders'),
                func.sum(Order.total_amount).label('revenue')
            ).filter(
                Order.created_at >= start_date
            ).group_by('year', 'week').order_by('year', 'week').all()
            
            data = [{
                'period': f"Week {trend.week}, {int(trend.year)}",
                'orders': trend.orders,
                'revenue': float(trend.revenue or 0)
            } for trend in trends]
            
        else:  # monthly
            # Group by month
            trends = db.session.query(
                extract('year', Order.created_at).label('year'),
                extract('month', Order.created_at).label('month'),
                func.count(Order.id).label('orders'),
                func.sum(Order.total_amount).label('revenue')
            ).filter(
                Order.created_at >= start_date
            ).group_by('year', 'month').order_by('year', 'month').all()
            
            data = [{
                'period': f"{int(trend.month)}/{int(trend.year)}",
                'orders': trend.orders,
                'revenue': float(trend.revenue or 0)
            } for trend in trends]
        
        return jsonify({
            'success': True,
            'period': period,
            'data': data
        })
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)}), 500

@report_bp.route('/api/top-customers', methods=['GET'])
@login_required
def top_customers():
    """Get top customers by revenue"""
    try:
        limit = request.args.get('limit', 10, type=int)
        days = request.args.get('days', 30, type=int)
        start_date = datetime.now() - timedelta(days=days)
        
        top_customers = db.session.query(
            Customer.id,
            Customer.name,
            Customer.email,
            func.count(Order.id).label('order_count'),
            func.sum(Order.total_amount).label('total_spent')
        ).join(Order).filter(
            Order.created_at >= start_date
        ).group_by(Customer.id, Customer.name, Customer.email).order_by(
            func.sum(Order.total_amount).desc()
        ).limit(limit).all()
        
        customers_data = [{
            'id': customer.id,
            'name': customer.name,
            'email': customer.email,
            'order_count': customer.order_count,
            'total_spent': float(customer.total_spent or 0)
        } for customer in top_customers]
        
        return jsonify({
            'success': True,
            'customers': customers_data
        })
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)}), 500

@report_bp.route('/api/top-products', methods=['GET'])
@login_required
def top_products():
    """Get top products by quantity sold"""
    try:
        limit = request.args.get('limit', 10, type=int)
        days = request.args.get('days', 30, type=int)
        start_date = datetime.now() - timedelta(days=days)
        
        top_products = db.session.query(
            Product.id,
            Product.name,
            Product.sku,
            func.sum(OrderItem.quantity).label('quantity_sold'),
            func.sum(OrderItem.total).label('revenue')
        ).join(OrderItem).join(Order).filter(
            Order.created_at >= start_date
        ).group_by(Product.id, Product.name, Product.sku).order_by(
            func.sum(OrderItem.quantity).desc()
        ).limit(limit).all()
        
        products_data = [{
            'id': product.id,
            'name': product.name,
            'sku': product.sku,
            'quantity_sold': int(product.quantity_sold or 0),
            'revenue': float(product.revenue or 0)
        } for product in top_products]
        
        return jsonify({
            'success': True,
            'products': products_data
        })
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)}), 500

@report_bp.route('/revenue-by-category', methods=['GET'])
@login_required
def revenue_by_category():
    """Get revenue breakdown by product category"""
    try:
        days = request.args.get('days', 30, type=int)
        start_date = datetime.now() - timedelta(days=days)
        
        category_revenue = db.session.query(
            Product.category,
            func.sum(OrderItem.total).label('revenue'),
            func.sum(OrderItem.quantity).label('quantity')
        ).join(OrderItem).join(Order).filter(
            Order.created_at >= start_date,
            Product.category.isnot(None),
            Product.category != ''
        ).group_by(Product.category).order_by(
            func.sum(OrderItem.total).desc()
        ).all()
        
        data = [{
            'category': cat or 'Uncategorized',
            'revenue': float(revenue or 0),
            'quantity': int(quantity or 0)
        } for cat, revenue, quantity in category_revenue]
        
        return jsonify({
            'success': True,
            'data': data
        })
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)}), 500

@report_bp.route('/customer-growth', methods=['GET'])
@login_required
def customer_growth():
    """Get customer growth over time"""
    try:
        days = request.args.get('days', 90, type=int)
        start_date = datetime.now() - timedelta(days=days)
        
        growth = db.session.query(
            func.date(Customer.created_at).label('date'),
            func.count(Customer.id).label('new_customers')
        ).filter(
            Customer.created_at >= start_date
        ).group_by(func.date(Customer.created_at)).order_by('date').all()
        
        data = [{
            'date': str(g.date),
            'new_customers': g.new_customers
        } for g in growth]
        
        return jsonify({
            'success': True,
            'data': data
        })
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)}), 500

@report_bp.route('/api/download', methods=['GET'])
@login_required
def download_report():
    """Download reports as PDF or Excel"""
    try:
        format_type = request.args.get('format', 'excel')  # 'excel' or 'pdf'
        report_type = request.args.get('type', 'summary')  # 'summary', 'customers', 'products', 'trends'
        days = request.args.get('days', 30, type=int)
        start_date = datetime.now() - timedelta(days=days)
        
        if format_type == 'excel' and OPENPYXL_AVAILABLE:
            wb = Workbook()
            
            if report_type == 'summary':
                ws = wb.active
                ws.title = "Sales Summary"
                
                # Get summary data
                total_revenue = db.session.query(func.sum(Order.total_amount)).filter(
                    Order.created_at >= start_date
                ).scalar() or 0.0
                total_orders = Order.query.filter(Order.created_at >= start_date).count()
                total_customers = Customer.query.count()
                active_customers = db.session.query(func.count(func.distinct(Order.customer_id))).filter(
                    Order.created_at >= start_date
                ).scalar() or 0
                
                # Write summary
                ws['A1'] = 'Sales Summary Report'
                ws['A1'].font = Font(bold=True, size=16)
                ws['A2'] = f'Period: Last {days} days'
                ws['A2'].font = Font(bold=True)
                
                ws['A4'] = 'Metric'
                ws['B4'] = 'Value'
                ws['A4'].font = Font(bold=True)
                ws['B4'].font = Font(bold=True)
                
                ws['A5'] = 'Total Revenue'
                ws['B5'] = float(total_revenue)
                ws['A6'] = 'Total Orders'
                ws['B6'] = total_orders
                ws['A7'] = 'Total Customers'
                ws['B7'] = total_customers
                ws['A8'] = 'Active Customers'
                ws['B8'] = active_customers
                
            elif report_type == 'customers':
                ws = wb.active
                ws.title = "Top Customers"
                
                limit = request.args.get('limit', 50, type=int)
                top = db.session.query(
                    Customer.id, Customer.name, Customer.email,
                    func.count(Order.id).label('order_count'),
                    func.sum(Order.total_amount).label('total_spent')
                ).join(Order).filter(Order.created_at >= start_date).group_by(
                    Customer.id, Customer.name, Customer.email
                ).order_by(func.sum(Order.total_amount).desc()).limit(limit).all()
                
                headers = ['Rank', 'Customer Name', 'Email', 'Orders', 'Total Spent']
                for col_num, header in enumerate(headers, 1):
                    cell = ws.cell(row=1, column=col_num, value=header)
                    cell.font = Font(bold=True)
                    cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
                    cell.font = Font(bold=True, color="FFFFFF")
                
                for row_num, customer in enumerate(top, 2):
                    ws.cell(row=row_num, column=1, value=row_num - 1)
                    ws.cell(row=row_num, column=2, value=customer.name)
                    ws.cell(row=row_num, column=3, value=customer.email)
                    ws.cell(row=row_num, column=4, value=customer.order_count)
                    ws.cell(row=row_num, column=5, value=float(customer.total_spent or 0))
                    
            elif report_type == 'products':
                ws = wb.active
                ws.title = "Top Products"
                
                limit = request.args.get('limit', 50, type=int)
                top = db.session.query(
                    Product.id, Product.name, Product.sku,
                    func.sum(OrderItem.quantity).label('quantity_sold'),
                    func.sum(OrderItem.total).label('revenue')
                ).join(OrderItem).join(Order).filter(
                    Order.created_at >= start_date
                ).group_by(Product.id, Product.name, Product.sku).order_by(
                    func.sum(OrderItem.quantity).desc()
                ).limit(limit).all()
                
                headers = ['Rank', 'Product Name', 'SKU', 'Quantity Sold', 'Revenue']
                for col_num, header in enumerate(headers, 1):
                    cell = ws.cell(row=1, column=col_num, value=header)
                    cell.font = Font(bold=True)
                    cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
                    cell.font = Font(bold=True, color="FFFFFF")
                
                for row_num, product in enumerate(top, 2):
                    ws.cell(row=row_num, column=1, value=row_num - 1)
                    ws.cell(row=row_num, column=2, value=product.name)
                    ws.cell(row=row_num, column=3, value=product.sku)
                    ws.cell(row=row_num, column=4, value=int(product.quantity_sold or 0))
                    ws.cell(row=row_num, column=5, value=float(product.revenue or 0))
            
            # Auto-adjust column widths
            for column in ws.columns:
                max_length = 0
                column_letter = column[0].column_letter
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                adjusted_width = min(max_length + 2, 50)
                ws.column_dimensions[column_letter].width = adjusted_width
            
            output = BytesIO()
            wb.save(output)
            output.seek(0)
            
            return send_file(
                output,
                mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
                as_attachment=True,
                download_name=f'report_{report_type}_{datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx'
            )
        else:
            # PDF export (if reportlab available)
            if REPORTLAB_AVAILABLE:
                buffer = BytesIO()
                doc = SimpleDocTemplate(buffer, pagesize=A4)
                elements = []
                styles = getSampleStyleSheet()
                
                # Title
                title = Paragraph(f"Sales Report - Last {days} days", styles['Title'])
                elements.append(title)
                elements.append(Spacer(1, 12))
                
                # Get summary data
                total_revenue = db.session.query(func.sum(Order.total_amount)).filter(
                    Order.created_at >= start_date
                ).scalar() or 0.0
                total_orders = Order.query.filter(Order.created_at >= start_date).count()
                
                # Summary table
                data = [['Metric', 'Value']]
                data.append(['Total Revenue', f'₹{total_revenue:,.2f}'])
                data.append(['Total Orders', str(total_orders)])
                
                table = Table(data)
                table.setStyle(TableStyle([
                    ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
                    ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                    ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
                    ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                    ('FONTSIZE', (0, 0), (-1, 0), 14),
                    ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
                    ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
                    ('GRID', (0, 0), (-1, -1), 1, colors.black)
                ]))
                elements.append(table)
                
                doc.build(elements)
                buffer.seek(0)
                
                return send_file(
                    buffer,
                    mimetype='application/pdf',
                    as_attachment=True,
                    download_name=f'report_{report_type}_{datetime.now().strftime("%Y%m%d_%H%M%S")}.pdf'
                )
            else:
                return jsonify({'success': False, 'error': 'PDF export not available'}), 500
                
    except Exception as e:
        import traceback
        traceback.print_exc()
        return jsonify({'success': False, 'error': str(e)}), 500
